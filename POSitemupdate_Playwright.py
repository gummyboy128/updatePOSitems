"""
OSPOS Item Update Automation Script
Automates the process of updating item prices in OSPOS using Playwright.
Reads Excel files containing item information and updates/creates items in the POS system.

Requirements:
    pip install playwright openpyxl coloredlogs pandas
    playwright install chromium
"""

from playwright.async_api import async_playwright, Page, Browser, BrowserContext, Playwright
from typing import List, Tuple, Optional, Dict
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import os
import logging
import coloredlogs
from datetime import datetime
import asyncio
from dataclasses import dataclass, field
import pandas as pd


# ==================== Configuration ====================
class Config:
    """Configuration settings for the automation script"""
    # Browser settings
    HEADLESS = False
    DEFAULT_TIMEOUT = 10000
    SCREENSHOT_DIR = "screenshots"

    # Login credentials
    USERNAME = "chee wei"
    PASSWORD = "heartwarmers"

    # URL options (uncomment the one you want to use)
    LOGIN_URL = "http://192.168.60.22/login"
    # LOGIN_URL = "http://88.88.88.88/login"
    # LOGIN_URL = "http://192.168.0.132/login"

    # Excel column indices (1-based)
    COL_BARCODE = 3
    COL_BRAND = 5
    COL_PRICE = 7
    COL_UNIT_PRICE = 8
    HEADER_ROW = 3
    DATA_START_ROW = 4

    # Default supplier when no filename match is found
    DEFAULT_SUPPLIER = 'Sheng Siong'

    # Logging
    DEBUG = True
    LOG_LEVEL = logging.INFO


# ==================== Supplier Mapping ====================
SUPPLIER_MAP: Dict[str, str] = {
    'ck': 'CK',
    'fruits_xs': 'SKC Trading',
    'ugroup': 'U-Group Holdings',
    'legacy': 'Legacy Food',
}


def get_supplier_name(file_name: str) -> str:
    """Resolve supplier name from Excel file name using SUPPLIER_MAP."""
    base = os.path.splitext(file_name)[0].lower()
    for keyword, supplier in SUPPLIER_MAP.items():
        if keyword in base:
            return supplier
    return Config.DEFAULT_SUPPLIER


# ==================== Custom Exceptions ====================
class OSPOSException(Exception):
    """Base exception for OSPOS automation errors"""
    pass


class ExcelValidationError(OSPOSException):
    """Raised when Excel file validation fails"""
    pass


class LoginError(OSPOSException):
    """Raised when login fails"""
    pass


class ItemUpdateError(OSPOSException):
    """Raised when item update fails"""
    pass


# ==================== Data Classes ====================
@dataclass
class ProductInfo:
    """Data class to hold product information"""
    name: str
    barcode: str
    price: float
    row_number: int

    @property
    def barcodes_list(self) -> List[str]:
        """Parse barcode string into list, handling multiple barcodes separated by '/'"""
        if "/" in str(self.barcode):
            return [bc.strip() for bc in str(self.barcode).split("/")]
        return [str(self.barcode)]


@dataclass
class ChangeRecord:
    """Record of a change made to an item"""
    item_name: str
    category: str
    barcode: str
    change_type: str  # 'price_update', 'category_update', 'supplier_update', 'new_item', 'name_update', 'error'
    old_value: str
    new_value: str
    details: str


@dataclass
class UpdateSummary:
    """Summary of updates performed on a worksheet"""
    worksheet_name: str
    updated_items: List[str]
    new_items: List[str]
    error_items: List[str]
    error_count: int
    change_records: List[ChangeRecord] = field(default_factory=list)


# ==================== Logging Setup ====================
class Logger:
    """Enhanced logging wrapper"""

    def __init__(self):
        """Set up file handler with timestamp-based log file and coloured console output."""
        log_filename = datetime.now().strftime('%Y_%m_%d_results.txt')
        logging.basicConfig(
            filename=log_filename,
            level=Config.LOG_LEVEL,
            filemode="w",
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        coloredlogs.install()
        self.logger = logging.getLogger(__name__)

    def info(self, msg: str) -> None:
        if Config.DEBUG:
            self.logger.info(f"[INFO] {msg}")

    def error(self, msg: str) -> None:
        self.logger.error(f"[ERROR] {msg}")

    def warning(self, msg: str) -> None:
        self.logger.warning(f"[WARNING] {msg}")

    def info_list(self, items: List[str]) -> None:
        for item in items:
            self.info(item)

    def error_list(self, items: List[str]) -> None:
        for item in items:
            self.error(item)


# Global logger instance
logger = Logger()


# ==================== Excel Processing ====================
class ExcelProcessor:
    """Handles Excel file reading and validation"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook: Optional[Workbook] = None

    def load_workbook(self) -> Workbook:
        """Load Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            logger.info(f"Opened stock list: {self.file_path}")
            return self.workbook
        except Exception as e:
            raise ExcelValidationError(f"Failed to load workbook: {e}")

    def validate_headers(self, worksheet: Worksheet) -> bool:
        """Validate price/unit-price column headers"""
        price_header = str(worksheet.cell(row=Config.HEADER_ROW, column=Config.COL_PRICE).value).lower()
        unit_price_header = str(worksheet.cell(row=Config.HEADER_ROW, column=Config.COL_UNIT_PRICE).value).lower()
        return "price" in price_header and "unit price" in unit_price_header

    def validate_item_headers(self, worksheet: Worksheet) -> bool:
        """Validate barcode and brand column headers"""
        barcode_header = str(worksheet.cell(row=Config.HEADER_ROW, column=Config.COL_BARCODE).value).lower()
        brand_header = str(worksheet.cell(row=Config.HEADER_ROW, column=Config.COL_BRAND).value).lower()
        return "barcode" in barcode_header and "brand" in brand_header

    def check_price_consistency(self) -> Tuple[bool, List[str]]:
        """
        Validate that 'Price' and 'Unit Price' columns match for all items.
        Returns: (is_valid, error_list)
        """
        if not self.workbook:
            self.load_workbook()

        error_list = []

        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]

            if not self.validate_headers(worksheet):
                continue

            logger.info(f"[{sheet_name}] Heading matches! Proceeding to check price...")

            current_row = Config.DATA_START_ROW
            while worksheet.cell(row=current_row, column=Config.COL_PRICE).value is not None:
                price = worksheet.cell(row=current_row, column=Config.COL_PRICE).value
                unit_price = worksheet.cell(row=current_row, column=Config.COL_UNIT_PRICE).value

                if price != unit_price:
                    item_name = worksheet.cell(row=current_row, column=Config.COL_BRAND).value
                    error_msg = f"{sheet_name} -- {item_name} (Row: {current_row})"
                    error_list.append(error_msg)
                    logger.error(f"Price mismatch: {error_msg}")

                current_row += 1

        if error_list:
            logger.error("SUMMARY: Price check FAILED")
            logger.error("SUMMARY: Please check the following items:")
            logger.error_list(error_list)
            return False, error_list

        logger.info("All price checks: PASS!")
        logger.info("#" * 65)
        return True, []


# ==================== Playwright Automation ====================
class OSPOSAutomation:
    """Main automation class for OSPOS item updates"""

    def __init__(self, supplier_name: str):
        self.supplier_name = supplier_name
        self.playwright: Optional[Playwright] = None
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None

    async def start_browser(self) -> None:
        """Initialize and start browser"""
        try:
            self.playwright = await async_playwright().start()
            self.browser = await self.playwright.chromium.launch(
                headless=Config.HEADLESS,
                args=["--start-maximized"]
            )
            self.context = await self.browser.new_context(no_viewport=True)
            self.page = await self.context.new_page()
            self.page.set_default_timeout(Config.DEFAULT_TIMEOUT)

            await self.page.goto(Config.LOGIN_URL, wait_until="networkidle")
            logger.info("Browser started successfully")
        except Exception as e:
            raise OSPOSException(f"Failed to start browser: {e}")

    async def close_browser(self) -> None:
        """Close browser and cleanup"""
        try:
            if self.context:
                await self.context.close()
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()
            logger.info("Browser closed successfully")
        except Exception as e:
            logger.error(f"Error closing browser: {e}")

    async def take_screenshot(self, step_name: str = "step") -> str:
        """Take and save screenshot"""
        os.makedirs(Config.SCREENSHOT_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"{Config.SCREENSHOT_DIR}/{step_name}_{timestamp}.png"
        await self.page.screenshot(path=filename)
        logger.info(f"Screenshot saved: {filename}")
        return filename

    # ── Auth & Navigation ──────────────────────────────────────────────────

    async def login(self) -> None:
        """Perform login to OSPOS"""
        try:
            username_input = self.page.locator("xpath=//input[@placeholder='Username']")
            await username_input.wait_for(state="visible")

            password_input = self.page.locator("xpath=//input[@placeholder='Password']")
            login_button = self.page.locator("xpath=//*[@name='login-button']")

            await username_input.fill(Config.USERNAME)
            await password_input.fill(Config.PASSWORD)
            await login_button.click()

            welcome_message = self.page.locator("text=Welcome to OSPOS")
            await welcome_message.wait_for(state="visible", timeout=10000)
            logger.info("Login successful")
        except Exception as e:
            await self.take_screenshot("login_error")
            raise LoginError(f"Login failed: {e}")

    async def navigate_to_items(self) -> None:
        """Navigate to Items tab"""
        try:
            items_menu = self.page.locator("xpath=//a[@title='Items']")
            await items_menu.wait_for(state="visible")
            await items_menu.click()

            search_bar = self.page.locator("xpath=//input[@placeholder='Search']")
            await search_bar.wait_for(state="visible")
            logger.info("Navigated to Items tab")
        except Exception as e:
            raise OSPOSException(f"Failed to navigate to Items tab: {e}")

    # ── Search & Table Helpers ─────────────────────────────────────────────

    async def search_item(self, search_term: str, timeout: int = 3000) -> bool:
        """
        Search for an item and wait for results.
        Returns True if records found, False if no records.
        """
        search_bar = self.page.locator("xpath=//input[@placeholder='Search']")
        await search_bar.fill("")
        await search_bar.fill(search_term)

        try:
            await self.page.wait_for_selector("xpath=//table[@id='table']/tbody/tr", timeout=timeout)
        except:
            pass

        await asyncio.sleep(2)

        content = await self.page.content()
        return "no-records-found" not in content

    async def get_table_row_count(self) -> int:
        """Get number of rows in results table"""
        rows = self.page.locator("xpath=//table[@id='table']/tbody/tr")
        await rows.first.wait_for(state="visible", timeout=5000)
        return await rows.count()

    async def _verify_success_message(self, message_text: str) -> None:
        """Wait for success message to appear"""
        try:
            success_message = self.page.locator(f"xpath=(//*[contains(text(), '{message_text}')])[1]")
            await success_message.wait_for(state="visible", timeout=20000)
        except Exception as e:
            logger.warning(f"Success message verification timeout: {e}")

    # ── Item Operations ────────────────────────────────────────────────────

    async def create_new_item(self, product: ProductInfo, category: str, barcode: str) -> Tuple[bool, Optional[ChangeRecord]]:
        """
        Create a new item in OSPOS.
        Returns: (success, change_record)
        """
        try:
            create_button = self.page.locator("xpath=//button[@title='New Item']")
            await create_button.wait_for(state="visible")
            await create_button.click()

            barcode_input = self.page.locator("xpath=//*[@name='item_number']")
            await barcode_input.wait_for(state="visible", timeout=5000)

            name_input = self.page.locator("xpath=//*[@name='name']")
            category_input = self.page.locator("xpath=//*[@name='category']")
            type_input = self.page.locator("xpath=//*[@name='stock_type'][@value='1']")
            supplier_input = self.page.locator("xpath=//*[@name='supplier_id']")
            wholesale_input = self.page.locator("xpath=//*[@name='cost_price']")
            retail_input = self.page.locator("xpath=//*[@name='unit_price']")

            await barcode_input.fill(barcode)
            await name_input.fill(product.name)
            await category_input.fill(category)
            await type_input.click()
            await supplier_input.select_option(label=self.supplier_name)

            await wholesale_input.clear()
            await retail_input.clear()
            await asyncio.sleep(0.3)
            await wholesale_input.fill(str(product.price))
            await asyncio.sleep(0.3)
            await retail_input.fill(str(product.price))

            close_button = self.page.locator("xpath=//button[@aria-label='close']")
            submit_button = self.page.locator("xpath=//button[@id='submit']")
            await submit_button.wait_for(state="visible")

            content = await self.page.content()
            if "form-group form-group-sm has-error" in content:
                logger.warning(f"Barcode already exists for: {product.name}")
                await close_button.click()
                await self.take_screenshot("barcode_error")
                return False, ChangeRecord(
                    item_name=product.name, category=category, barcode=barcode,
                    change_type='error', old_value='N/A', new_value='N/A',
                    details=f"Failed to create - Barcode {barcode} already exists"
                )

            await submit_button.click()
            await self._verify_success_message(f"You have successfully added item {product.name}")
            logger.info(f"Created new item: {product.name}")

            return True, ChangeRecord(
                item_name=product.name, category=category, barcode=barcode,
                change_type='new_item', old_value='N/A', new_value=f"${product.price}",
                details=f"Created new item with barcode {barcode} at ${product.price}"
            )

        except Exception as e:
            logger.error(f"Failed to create item {product.name}: {e}")
            await self.take_screenshot("create_item_error")
            return False, ChangeRecord(
                item_name=product.name, category=category, barcode=barcode,
                change_type='error', old_value='N/A', new_value='N/A',
                details=f"Failed to create - Error: {str(e)}"
            )

    async def update_item_name(self, product: ProductInfo) -> None:
        """Update existing item name"""
        try:
            update_button = self.page.locator("xpath=(//*[@title='Update Item'])[1]")
            await update_button.wait_for(state="visible")
            await update_button.click()

            name_input = self.page.locator("xpath=//*[@name='name']")
            await name_input.wait_for(state="visible", timeout=5000)

            submit_button = self.page.locator("xpath=//button[@id='submit']")
            await name_input.clear()
            await name_input.fill(product.name)
            await submit_button.click()

            await self._verify_success_message("You have successfully updated item")
            logger.info(f"Updated item name: {product.name}")
        except Exception as e:
            logger.error(f"Failed to update item name {product.name}: {e}")

    async def update_item_prices(self, product: ProductInfo, category: str) -> Tuple[List[str], List[ChangeRecord]]:
        """
        Update price, category, and supplier for all matching items in the table.
        Returns: (updated_item_names, change_records)
        """
        updated_items = []
        change_records = []

        try:
            row_count = await self.get_table_row_count()

            for row_index in range(1, row_count + 1):
                category_cell = await self.page.locator(f"xpath=//tbody/tr[{row_index}]/td[5]").text_content()
                current_category = category_cell.strip()
                category_matches = current_category.lower() == category.lower()

                wholesale_cell = await self.page.locator(f"xpath=//tbody/tr[{row_index}]/td[7]").text_content()
                current_wholesale = float(wholesale_cell.strip("$"))

                barcode_cell = await self.page.locator(f"xpath=//tbody/tr[{row_index}]/td[3]").text_content()
                current_barcode = barcode_cell.strip()

                needs_update = current_wholesale != product.price or not category_matches

                if not needs_update:
                    continue

                update_button = self.page.locator(f"xpath=(//*[@title='Update Item'])[{row_index}]")
                await update_button.wait_for(state="visible")
                await update_button.click()

                wholesale_input = self.page.locator("xpath=//*[@name='cost_price']")
                await wholesale_input.wait_for(state="visible", timeout=5000)

                category_input = self.page.locator("xpath=//*[@name='category']")
                supplier_input = self.page.locator("xpath=//*[@name='supplier_id']")
                retail_input = self.page.locator("xpath=//*[@name='unit_price']")
                submit_button = self.page.locator("xpath=//button[@id='submit']")

                if not category_matches:
                    await category_input.clear()
                    await category_input.fill(category)
                    logger.info(f"Updating category from '{current_category}' to '{category}' for: {product.name}")
                    change_records.append(ChangeRecord(
                        item_name=product.name, category=category, barcode=current_barcode,
                        change_type='category_update', old_value=current_category, new_value=category,
                        details=f"Category changed from '{current_category}' to '{category}'"
                    ))

                # Always ensure supplier is correct
                current_supplier_option = await supplier_input.evaluate("el => el.options[el.selectedIndex]?.text ?? ''")
                if current_supplier_option.strip() != self.supplier_name:
                    await supplier_input.select_option(label=self.supplier_name)
                    logger.info(f"Updating supplier from '{current_supplier_option.strip()}' to '{self.supplier_name}' for: {product.name}")
                    change_records.append(ChangeRecord(
                        item_name=product.name, category=category, barcode=current_barcode,
                        change_type='supplier_update',
                        old_value=current_supplier_option.strip(),
                        new_value=self.supplier_name,
                        details=f"Supplier changed from '{current_supplier_option.strip()}' to '{self.supplier_name}'"
                    ))

                if current_wholesale != product.price:
                    await wholesale_input.clear()
                    await retail_input.clear()
                    await asyncio.sleep(0.3)
                    await wholesale_input.fill(str(product.price))
                    await asyncio.sleep(0.3)
                    await retail_input.fill(str(product.price))
                    logger.info(f"Updating price from ${current_wholesale} to ${product.price} for: {product.name}")
                    change_records.append(ChangeRecord(
                        item_name=product.name, category=category, barcode=current_barcode,
                        change_type='price_update',
                        old_value=f"${current_wholesale}", new_value=f"${product.price}",
                        details=f"Price changed from ${current_wholesale} to ${product.price}"
                    ))

                await submit_button.click()
                await self._verify_success_message("You have successfully updated item")
                updated_items.append(product.name)

        except Exception as e:
            logger.error(f"Failed to update item {product.name}: {e}")

        return updated_items, change_records

    # ── Worksheet / File Processing ────────────────────────────────────────

    async def process_excel_file(self, file_path: str) -> List[UpdateSummary]:
        """Process all worksheets in an Excel file and return summaries"""
        excel_processor = ExcelProcessor(file_path)
        workbook = excel_processor.load_workbook()

        logger.info("Starting to update OSPOS...")
        all_summaries = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            if not excel_processor.validate_item_headers(worksheet):
                logger.warning(f"[{sheet_name}] Headers do not match, skipping worksheet")
                continue

            logger.info(f"[Worksheet] Working on: {sheet_name}")
            summary = await self._process_worksheet(worksheet, sheet_name.strip())
            self._print_summary(summary)
            all_summaries.append(summary)

        return all_summaries

    async def _process_worksheet(self, worksheet: Worksheet, sheet_name: str) -> UpdateSummary:
        """Process a single worksheet"""
        summary = UpdateSummary(
            worksheet_name=sheet_name,
            updated_items=[],
            new_items=[],
            error_items=[],
            error_count=0
        )

        current_row = Config.DATA_START_ROW
        while worksheet.cell(row=current_row, column=Config.COL_BRAND).value is not None:
            cell_value = worksheet.cell(row=current_row, column=Config.COL_BRAND).value
            if cell_value == "Total":
                break

            product = ProductInfo(
                name=str(cell_value).strip(),
                barcode=worksheet.cell(row=current_row, column=Config.COL_BARCODE).value,
                price=float(worksheet.cell(row=current_row, column=Config.COL_PRICE).value),
                row_number=current_row
            )

            await self._process_product(product, sheet_name, summary)
            current_row += 1

        return summary

    async def _process_product(self, product: ProductInfo, category: str, summary: UpdateSummary) -> None:
        """Search, update, or create a single product"""
        found_by_name = await self.search_item(product.name)

        if not found_by_name:
            found_by_barcode = False
            for barcode in product.barcodes_list:
                if await self.search_item(barcode):
                    found_by_barcode = True
                    await self.update_item_name(product)
                    updated_items, change_records = await self.update_item_prices(product, category)
                    summary.updated_items.extend(updated_items)
                    summary.change_records.extend(change_records)
                    break

            if not found_by_barcode:
                logger.info(f"No record found, creating new item: {product.name}")
                for barcode in product.barcodes_list:
                    success, change_record = await self.create_new_item(product, category, barcode)
                    if change_record:
                        summary.change_records.append(change_record)
                    if success:
                        summary.new_items.append(product.name)
                    else:
                        summary.error_items.append(product.name)
                        summary.error_count += 1
        else:
            row_count = await self.get_table_row_count()

            if row_count >= len(product.barcodes_list):
                # Exact match or more records in system — update existing
                updated_items, change_records = await self.update_item_prices(product, category)
                summary.updated_items.extend(updated_items)
                summary.change_records.extend(change_records)
            else:
                # Excel has more barcodes — create the missing ones
                for barcode in product.barcodes_list:
                    if not await self.search_item(barcode):
                        logger.info(f"Existing item with new barcode, creating: {product.name}")
                        success, change_record = await self.create_new_item(product, category, barcode)
                        if change_record:
                            summary.change_records.append(change_record)
                        if success:
                            summary.new_items.append(product.name)
                        else:
                            summary.error_items.append(product.name)
                            summary.error_count += 1

    def _print_summary(self, summary: UpdateSummary) -> None:
        """Print summary of worksheet processing"""
        logger.info("-" * 55)
        logger.info(f"[{summary.worksheet_name}] - Summary:")
        logger.info(f"[{summary.worksheet_name}] - Items updated:")
        logger.info_list(summary.updated_items if summary.updated_items else ["None"])
        logger.info(f"[{summary.worksheet_name}] - Items added:")
        logger.info_list(summary.new_items if summary.new_items else ["None"])
        logger.info(f"[{summary.worksheet_name}] - Items with errors:")
        logger.error_list(summary.error_items if summary.error_items else ["None"])
        logger.info("-" * 55)

    # ── Report Generation ──────────────────────────────────────────────────

    @staticmethod
    def generate_detailed_report(summaries: List[UpdateSummary], filename: str) -> str:
        """
        Generate detailed Excel report with all changes.
        Returns path to generated report, or empty string on failure.
        """
        timestamp = datetime.now().strftime('%Y_%m_%d_%H%M%S')
        report_path = f"reports/{filename}_{timestamp}_detailed_report.xlsx"

        all_changes = [record for summary in summaries for record in summary.change_records]

        if not all_changes:
            logger.warning("No changes to report")
            return ""

        changes_data = [
            {
                'Item Name': r.item_name,
                'Category': r.category,
                'Barcode': r.barcode,
                'Change Type': r.change_type.replace('_', ' ').title(),
                'Old Value': r.old_value,
                'New Value': r.new_value,
                'Details': r.details,
            }
            for r in all_changes
        ]
        df_changes = pd.DataFrame(changes_data)

        summary_data = [
            {
                'Worksheet': s.worksheet_name,
                'Items Updated': len(set(s.updated_items)),
                'Items Added': len(set(s.new_items)),
                'Errors': s.error_count,
                'Total Changes': len(s.change_records),
            }
            for s in summaries
        ]
        df_summary = pd.DataFrame(summary_data)
        change_types_data = df_changes.groupby('Change Type').size().reset_index(name='Count')

        try:
            os.makedirs("reports", exist_ok=True)
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                df_changes.to_excel(writer, sheet_name='All Changes', index=False)
                change_types_data.to_excel(writer, sheet_name='Change Types', index=False)

                for change_type in df_changes['Change Type'].unique():
                    df_type = df_changes[df_changes['Change Type'] == change_type]
                    df_type.to_excel(writer, sheet_name=change_type[:31], index=False)

                for sheet_name, worksheet in writer.sheets.items():
                    for column in worksheet.columns:
                        max_length = max(
                            (len(str(cell.value)) for cell in column if cell.value),
                            default=0
                        )
                        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

            logger.info(f"Detailed report generated: {report_path}")
            return report_path

        except Exception as e:
            logger.error(f"Failed to generate report: {e}")
            return ""


# ==================== Main Execution ====================
async def main() -> None:
    """Main execution function"""
    cwd = os.getcwd()
    files_processed = 0

    logger.info("=" * 65)
    logger.info("OSPOS Item Update Automation Started")
    logger.info("=" * 65)

    excel_files = [
        f for f in os.listdir(cwd)
        if os.path.isfile(os.path.join(cwd, f))
        and f.endswith('.xlsx')
        and not f.startswith('~$')
        and '_detailed_report.xlsx' not in f
    ]

    if not excel_files:
        logger.error("No Excel files found in current directory!")
        return

    logger.info(f"Found {len(excel_files)} Excel file(s) to process")

    for file_name in excel_files:
        file_path = os.path.join(cwd, file_name)
        supplier_name = get_supplier_name(file_name)

        logger.info(f"\nProcessing file: {file_name}")
        logger.info(f"Supplier: {supplier_name}")

        try:
            excel_processor = ExcelProcessor(file_path)
            is_valid, _ = excel_processor.check_price_consistency()

            if not is_valid:
                logger.error(f"Price validation failed for {file_name}")
                logger.error("Please check with supplier on the price mismatch for 'Price' and 'Unit Price'")
                continue

            automation = OSPOSAutomation(supplier_name=supplier_name)
            await automation.start_browser()
            await automation.login()
            await automation.navigate_to_items()
            summaries = await automation.process_excel_file(file_path)
            await automation.close_browser()

            if summaries:
                base_filename = os.path.splitext(file_name)[0]
                report_path = OSPOSAutomation.generate_detailed_report(summaries, base_filename)
                if report_path:
                    logger.info(f"Detailed report saved: {report_path}")

            files_processed += 1
            logger.info(f"Successfully processed: {file_name}")

        except ExcelValidationError as e:
            logger.error(f"Excel validation error for {file_name}: {e}")
        except LoginError as e:
            logger.error(f"Login error: {e}")
        except OSPOSException as e:
            logger.error(f"OSPOS automation error for {file_name}: {e}")
        except Exception as e:
            logger.error(f"Unexpected error processing {file_name}: {e}")
            import traceback
            logger.error(traceback.format_exc())

    logger.info("=" * 65)
    if files_processed > 0:
        logger.info(f"FINISHED PROCESSING - {files_processed} file(s) completed")
    else:
        logger.error("Zero files processed successfully!")
    logger.info("=" * 65)


if __name__ == "__main__":
    asyncio.run(main())
